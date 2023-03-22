
"""
Created on Wed Mar 22 10:29:08 2023

@author: Rehan
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

# Fetch data from MITRE ATT&CK Enterprise Framework API
def fetch_mitre_data():
    url = "https://raw.githubusercontent.com/mitre/cti/master/enterprise-attack/enterprise-attack.json"
    response = requests.get(url)
    return response.json()

# Process fetched data into a DataFrame
def process_data(raw_data):
    tactics = {}
    techniques = []
    threat_actors = []

    # First, process all tactics
    for item in raw_data['objects']:
        if item['type'] == 'x-mitre-tactic':
            tactics[item['x_mitre_shortname']] = item['name']

    # Then, process all techniques
    for item in raw_data['objects']:
        if item['type'] == 'attack-pattern':
            for phase in item['kill_chain_phases']:
                if phase['kill_chain_name'] == 'mitre-attack' and phase['phase_name'] in tactics:
                    technique = {
                        'tactic_id': phase['phase_name'],
                        'tactic': tactics[phase['phase_name']],
                        'technique_id': item['external_references'][0]['external_id'],
                        'technique': item['name']
                    }
                    techniques.append(technique)

        # Process threat actors
        if item['type'] == 'intrusion-set':
            threat_actor = {
                'threat_actor_id': item['external_references'][0]['external_id'],
                'threat_actor': item['name']
            }
            threat_actors.append(threat_actor)

    df_techniques = pd.DataFrame(techniques)
    df_threat_actors = pd.DataFrame(threat_actors)

    return df_techniques, df_threat_actors, tactics


# Save DataFrame to Excel
def save_to_excel(df, sheet_name, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Set header styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
        if ws.row_dimensions[ws.max_row].height is None:
            ws.row_dimensions[ws.max_row].height = 45
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.alignment = header_alignment

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
                ws.column_dimensions[column].width = max_length + 2

    wb.save(filename)

def filter_techniques_by_threat_actor(raw_data, df_techniques, df_threat_actors, tactics, threat_actor_name):
    threat_actor = df_threat_actors[df_threat_actors['threat_actor'].str.contains(threat_actor_name, case=False)]

    if not threat_actor.empty:
        threat_actor_id = threat_actor.iloc[0]['threat_actor_id']

        technique_ids = set()
        for item in raw_data['objects']:
            if item['type'] == 'intrusion-set' and item['external_references'][0]['external_id'] == threat_actor_id:
                threat_actor_full_id = item['id']
                break

        for item in raw_data['objects']:
            if item['type'] == 'relationship' and item['source_ref'] == threat_actor_full_id and item['target_ref'].startswith('attack-pattern--'):
                technique_id = item['target_ref']
                technique_ids.add(technique_id)

        relevant_techniques = []
        for technique_id in technique_ids:
            for technique in raw_data['objects']:
                if technique['id'] == technique_id:
                    external_id = None
                    for reference in technique['external_references']:
                        if reference['source_name'] == 'mitre-attack':
                            external_id = reference['external_id']
                            break

                    if external_id:
                        for phase in technique['kill_chain_phases']:
                            if phase['kill_chain_name'] == 'mitre-attack' and phase['phase_name'] in tactics:
                                relevant_technique = {
                                    'tactic_id': phase['phase_name'],
                                    'tactic': tactics[phase['phase_name']],
                                    'technique_id': external_id,
                                    'technique': technique['name']
                                }
                                relevant_techniques.append(relevant_technique)
                                break

        return pd.DataFrame(relevant_techniques)
    else:
        print(f"No threat actor found with the name: {threat_actor_name}")
        return None
    
if __name__ == "__main__":
    raw_data = fetch_mitre_data()
    df_techniques, df_threat_actors, tactics = process_data(raw_data)

    threat_actor_name = "OilRig"  # Replace this with the desired threat actor name
    filtered_techniques = filter_techniques_by_threat_actor(raw_data, df_techniques, df_threat_actors, tactics, threat_actor_name)

    if filtered_techniques is not None:
        save_to_excel(filtered_techniques, 'Filtered_Techniques', f'{threat_actor_name}_techniques.xlsx')

